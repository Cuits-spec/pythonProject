import openpyxl

# 打开工作簿
workbook = openpyxl.load_workbook('../test_api/Excelread/test_api.xlsx')
ws = workbook.active
# 获取表单
sheet = workbook['Sheet1']
# 读取指定数据
cell = sheet.cell(row=2, column=1)

# a = cell(row=2, column=1).value
# print(cell)
# print(workbook.sheetnames)

cell_A2 = ws['A2']
# print(cell_A2.value)


import openpyxl

class ExcelHelper:
    def __int__(self,file_path):
        self.file_path = file_path
        self.workbook = None

    def open_excel(self):
        # 打开excel
        workbook = openpyxl.load_workbook(self.file_path)
        self.workbook = workbook
        return workbook

    def get_sheet(self,sheet_name):
        # 获取sheet
        workbook = self.open_excel()
        sheet = workbook[sheet_name]
        return sheet

    def read_excel(self,sheet_name):
        sheet = self.get_sheet(sheet_name)
        rows = list(sheet.rows)












def read_data(finename, sheetname):
    # 读取Excel
    wb = openpyxl.load_workbook(finename)
    # 读取表单
    sheet = wb[sheetname]
    max_row = sheet.max_row
    # print('行数：', max_row)
    max_column = sheet.max_column
    # print('列数：', max_column)
    # 读取最大行数
    max_row = sheet.max_row
    # 创建一个空列表存放用例
    list = []
    for i in range(2, max_row +1):
        dict1 = dict(url=sheet.cell(row=i, column=2).value,
                     method=sheet.cell(row=i,column=3).value,
                     headers=sheet.cell(row=i,column=4).value,
                     data=sheet.cell(row=i, column=5).value,
                     status_code=sheet.cell(row=i,column=6).value,
                     expect=sheet.cell(row=i, column=7).value
                     )
        list.append(dict1)
    return list



    # list1 = []
    # for x in range(2, max_row-18):
    #     list1 = []
    #     for y in range(2, max_column + 1):
    #         list1.append(sheet.cell(row=x, column=y).value)
    #     list1.append(list1)
    #
    # return list1





#
cases = read_data('../test_api/Excelread/test_api.xlsx', 'Sheet2')
print(cases)