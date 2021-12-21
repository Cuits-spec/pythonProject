import openpyxl

class Read_Excel():

    def read_excel(self,filename, sheetname):
        wb = openpyxl.load_workbook(filename)
        # print(wb.sheetnames)
        ws = wb[sheetname]
        # print(ws['B1'].value)
        max_row = ws.max_row
        print('行数：',max_row)
        max_column = ws.max_column
        print('列数：',max_column)
        test_data = []
        for x in range(2, max_row -18):
            test_case = []
            for y in range(2, max_column + 1):
                test_case.append(ws.cell(row=x, column=y).value)
                # print(ws.cell(row=x,column=y).value)
            test_data.append(test_case)

        return test_data

read1 = Read_Excel()

excel =read1.read_excel('../../pytest_api/data/test_api.xlsx', 'Sheet1')
print(excel)


# class Excel:
#     def excel(self,path):
#         excel = openpyxl.load_workbook(path)
#         sheet = excel['Sheet1']
#         max_row = sheet.max_row
#         print('行数：',max_row)
#         max_column = sheet.max_column
#         print('列数：',max_column)
#         data_list = []
#
#         for row in range(2,max_row-19):
#             row_list = []
#             for column in range(2,max_column+1):
#                 # 获取单元格值
#                 value = sheet.cell(row,column).value
#                 row_list.append(value)
#             data_list.append(row_list)
#         return data_list
#
# if __name__ == '__main__':
#     print(Excel().excel('../../pytest_api/data/test_api.xlsx'))
#





