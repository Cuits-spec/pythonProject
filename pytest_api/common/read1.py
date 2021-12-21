import openpyxl


def read_data(finename, sheetname):
    # 读取Excel
    wb = openpyxl.load_workbook(finename)
    # 读取表单
    sheet = wb[sheetname]
    max_row = sheet.max_row
    # print('行数：', max_row)
    max_column = sheet.max_column
    # print('列数：', max_column)
    # 读取最大行数
    max_row = sheet.max_row
    # 创建一个空列表存放用例
    list = []
    for i in range(2, max_row - 18):
        dict1 = dict(url=sheet.cell(row=i, column=1).value,
                     method=sheet.cell(row=i,column=2).value,
                     data=sheet.cell(row=i, column=3).value,
                     status_code=sheet.cell(row=i,column=4).value,
                     expect=sheet.cell(row=i, column=5).value
                     )
        list.append(dict1)
    return list


# cases = read_data('../../pytest_api/data/test_api.xlsx', 'Sheet1')
# print(cases)