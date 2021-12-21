import openpyxl
import requests
from test_api.common import api


# 表名filename；
# 表的sheet名sheetname；
# 表的单元格cell的坐标row及column；

def read_data(finename, sheetname):
    # 读取Excel
    wb = openpyxl.load_workbook(finename)
    # 读取表单
    sheet = wb[sheetname]
    # 读取最大行数
    max_row = sheet.max_row
    # 创建一个空列表存放用例
    list = []
    for i in range(2, max_row + 1):
        dict1 = dict(case_name=sheet.cell(row=i, column=1).value,
                     url=sheet.cell(row=i, column=2).value,
                     header=sheet.cell(row=i, column=4).value,
                     data=sheet.cell(row=i, column=5).value,
                     expect=sheet.cell(row=i, column=7).value
                     )
        list.append(dict1)
    return list


cases = read_data('test_api.xlsx', 'Sheet1')
# print(cases)
a = api.HttpRequest()
for case in cases:
    url = case.get('url')
    data = eval(case.get('data'))
    res = a.http_request(url, data, 'post')

    print(res.json())


# def test_read():
#     cases = read_data('test_api.xlsx', 'Sheet1')
#     # print(cases)
#     a = api.HttpRequest()
#     for case in cases:
#         url = case.get('url')
#         data = eval(case.get('data'))
#         res = a.http_request(url, data, 'post')
#
#         print(res.json())
#
# test_read()