import openpyxl
import pytest
import requests
import json
from pytest_api.common import read

# wb = openpyxl.load_workbook('test_api.xlsx')
# # print(wb.sheetnames)
# ws = wb['Sheet1']
# # print(ws['B1'].value)
# test_data = []
# for x in range(2, len(tuple(ws.rows))+1):
#     test_case = []
#     for y in range(2,7):
#         test_case.append(ws.cell(row=x,column=y).value)
#         # print(ws.cell(row=x,column=y).value)
#     test_data.append(test_case)
# print(test_data)

sss = read.read_excel('test_api.xlsx', 'Sheet1')
print(sss)




@pytest.mark.parametrize('url,method,data,code,Expect_res' , sss)
def test_excel(url, method,data,code,Expect_res):
    if method == 'POST':
        res = requests.post(url, data=json.load(data))
        assert res.status_code == code
        assert res.json() == json.load(Expect_res)
